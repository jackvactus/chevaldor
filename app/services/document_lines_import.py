"""Import Excel des lignes devis / facture."""
from __future__ import annotations

import re
from io import BytesIO
from typing import Any, Optional

import openpyxl

_HEADER_MAP = {
    "description": ("description", "désignation", "designation", "libellé", "libelle", "prestation", "produit"),
    "quantity": ("qté", "qte", "quantité", "quantite", "qty", "quantity"),
    "unit_price": ("pu", "p.u.", "prix", "prix unitaire", "pu ht", "montant unitaire", "unit price"),
    "discount_pct": ("remise", "rem.", "remise %", "discount", "rem %"),
    "vat_rate": ("tva", "tva %", "tva%", "vat"),
}


def _parse_amount(raw: Any) -> float:
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().upper().replace("FCFA", "").replace("F", "").replace(" ", "")
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def _norm_header(val: Any) -> str:
    return re.sub(r"\s+", " ", str(val or "").strip().lower())


def _detect_columns(headers: list[Any]) -> dict[str, int]:
    cols: dict[str, int] = {}
    for idx, h in enumerate(headers):
        nh = _norm_header(h)
        if not nh:
            continue
        for key, aliases in _HEADER_MAP.items():
            if key in cols:
                continue
            if nh in aliases or any(a in nh for a in aliases):
                cols[key] = idx
    if "description" not in cols and len(headers) >= 1:
        cols["description"] = 0
    if "quantity" not in cols and len(headers) >= 2:
        cols["quantity"] = 1
    if "unit_price" not in cols and len(headers) >= 3:
        cols["unit_price"] = 2
    return cols


def parse_document_lines_excel(content: bytes, filename: str = "", default_vat_rate: float = 0.0) -> dict[str, Any]:
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row < 2:
        return {"ok": False, "message": "Fichier vide ou sans données", "lines": []}

    header_row = 1
    for r in range(1, min(6, max_row + 1)):
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        joined = " ".join(_norm_header(v) for v in row_vals if v)
        if any(k in joined for k in ("description", "désignation", "designation", "qté", "qte", "pu")):
            header_row = r
            break

    headers = [ws.cell(header_row, c).value for c in range(1, max_col + 1)]
    cols = _detect_columns(headers)
    lines: list[dict] = []
    errors: list[str] = []

    for row_idx in range(header_row + 1, max_row + 1):
        desc = ws.cell(row_idx, cols.get("description", 0) + 1).value
        desc_s = str(desc or "").strip()
        qty = _parse_amount(ws.cell(row_idx, cols["quantity"] + 1).value) if "quantity" in cols else 1.0
        price = _parse_amount(ws.cell(row_idx, cols["unit_price"] + 1).value) if "unit_price" in cols else 0.0
        disc = _parse_amount(ws.cell(row_idx, cols["discount_pct"] + 1).value) if "discount_pct" in cols else 0.0
        vat = _parse_amount(ws.cell(row_idx, cols["vat_rate"] + 1).value) if "vat_rate" in cols else default_vat_rate

        if not desc_s and price <= 0:
            continue
        if not desc_s:
            errors.append(f"Ligne {row_idx} : description manquante")
            continue
        if qty <= 0:
            qty = 1.0
        lines.append({
            "description": desc_s,
            "quantity": qty,
            "unit_price": price,
            "discount_pct": disc,
            "vat_rate": vat if vat > 0 else default_vat_rate,
            "product_type": "SERVICE",
            "position": len(lines),
        })

    return {
        "ok": bool(lines),
        "filename": filename,
        "sheet": ws.title,
        "header_row": header_row,
        "lines": lines,
        "imported": len(lines),
        "errors": errors,
        "message": f"{len(lines)} ligne(s) détectée(s)" if lines else "Aucune ligne valide",
    }
