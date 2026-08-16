"""Import FICHE DE COLLECTE.xlsx → plans factures récurrentes (collecte journalière)."""
from __future__ import annotations

import re
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any, Optional

import openpyxl
from sqlalchemy.orm import Session

from app.models import Client
from app.models_business_ext import CollectionPayment
from app.models_recurring_advanced import PaymentRecurrence


def _now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S")

CLIENTELE_SEGMENT = "Clientèle PC"
MONTANT_COL = 2
DATA_START_ROW = 4
HEADER_ROW_DATES = 2
HEADER_ROW_LABELS = 3


def _parse_amount(raw: Any) -> float:
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().upper()
    s = s.replace("FCFA", "").replace("F", "").replace(" ", "").replace("\u00a0", "")
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def _as_date(val: Any) -> Optional[date]:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _norm_name(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip().upper())


def parse_collecte_workbook(content: bytes, filename: str = "", *, payments_limit: Optional[int] = 30) -> dict[str, Any]:
    """Parse la fiche de collecte (matrice dates × clients)."""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    max_col = ws.max_column or 0
    max_row = ws.max_row or 0

    date_cols: list[tuple[int, date]] = []
    for col in range(3, max_col + 1):
        d = _as_date(ws.cell(HEADER_ROW_DATES, col).value)
        if d:
            date_cols.append((col, d))

    if not date_cols:
        for col in range(2, max_col + 1):
            d = _as_date(ws.cell(HEADER_ROW_DATES, col).value)
            if d and col != MONTANT_COL:
                date_cols.append((col, d))

    rows: list[dict] = []
    for row_idx in range(DATA_START_ROW, max_row + 1):
        name_raw = ws.cell(row_idx, 1).value
        if not name_raw or not str(name_raw).strip():
            continue
        name = str(name_raw).strip()
        low = name.lower()
        if low in ("nom &prénom", "nom & prenom", "nom", "total", "totaux"):
            continue

        montant_total = _parse_amount(ws.cell(row_idx, MONTANT_COL).value)
        payments: list[dict] = []
        for col, d in date_cols:
            amt = _parse_amount(ws.cell(row_idx, col).value)
            if amt > 0:
                payments.append({"date": d.isoformat(), "amount": amt})

        if montant_total <= 0 and not payments:
            continue

        daily_amounts = [p["amount"] for p in payments]
        typical = 0.0
        if daily_amounts:
            typical = round(sum(daily_amounts) / len(daily_amounts), 2)
        elif montant_total > 0:
            typical = montant_total

        rows.append({
            "name": name,
            "montant_total": montant_total,
            "typical_daily": typical,
            "payments_count": len(payments),
            "payments_total": round(sum(daily_amounts), 2),
            "first_date": payments[0]["date"] if payments else None,
            "last_date": payments[-1]["date"] if payments else None,
            "payments": payments if payments_limit is None else payments[:payments_limit],
        })

    return {
        "ok": True,
        "filename": filename,
        "sheet": ws.title,
        "dates": [d.isoformat() for _, d in date_cols],
        "rows": rows,
        "total": len(rows),
    }


def _find_client(db: Session, name: str) -> Optional[Client]:
    norm = _norm_name(name)
    clients = db.query(Client).filter(Client.segment == CLIENTELE_SEGMENT).all()
    for c in clients:
        if _norm_name(c.name) == norm:
            return c
    for c in clients:
        if norm in _norm_name(c.name) or _norm_name(c.name) in norm:
            return c
    return None


def _ensure_client(db: Session, name: str, *, create_missing: bool, company_id: Optional[int] = None) -> tuple[Optional[Client], bool]:
    existing = _find_client(db, name)
    if existing:
        return existing, False
    if not create_missing:
        return None, False
    parts = name.split(" ", 1)
    client = Client(
        name=name.strip(),
        type="Particulier",
        segment=CLIENTELE_SEGMENT,
        contact=parts[0],
        status="actif",
        notes="Importé depuis FICHE DE COLLECTE",
    )
    if company_id is not None:
        client.company_id = company_id
    db.add(client)
    db.flush()
    return client, True


def import_collecte_to_recurring(
    db: Session,
    content: bytes,
    *,
    filename: str = "",
    create_clients: bool = True,
    frequency: str = "daily",
    mode: str = "merge",
    company_id: Optional[int] = None,
) -> dict[str, Any]:
    """
    Crée ou met à jour des plans récurrents depuis la fiche.
    mode: merge | replace (replace désactive les plans collecte existants du même libellé)
    """
    parsed = parse_collecte_workbook(content, filename, payments_limit=None)
    if not parsed.get("rows"):
        return {
            "ok": False,
            "message": "Aucune ligne client exploitable dans le fichier",
            "parsed": parsed,
        }

    created = updated = skipped = clients_created = 0
    real_payments_created = real_payments_skipped = 0
    errors: list[str] = []

    if mode == "replace":
        for plan in db.query(PaymentRecurrence).filter(
            PaymentRecurrence.category == "collecte"
        ).all():
            plan.status = 'cancelled'
            plan.is_active = False

    for row in parsed["rows"]:
        name = row["name"]
        client, was_new = _ensure_client(db, name, create_missing=create_clients, company_id=company_id)
        if not client:
            skipped += 1
            errors.append(f"{name} : client introuvable")
            continue
        if was_new:
            clients_created += 1

        # Paiements réels (un par date de la matrice) — persistés dans collection_payments, la
        # table lue par le calendrier « montants réels » (_unified_collection_rows). Le plan
        # PaymentRecurrence créé/mis à jour plus bas ne sert lui qu'à projeter les échéances
        # futures ; avant ce correctif, l'historique parsé ici (row["payments"]) était calculé
        # (moyenne "typical_daily") puis jeté sans jamais être enregistré nulle part de lisible
        # par le calendrier — d'où un calendrier "réel" toujours vide malgré un import "réussi".
        existing_dates = {
            (p_date, round(float(p_amt or 0), 2))
            for p_date, p_amt in db.query(
                CollectionPayment.collection_date, CollectionPayment.amount
            ).filter(CollectionPayment.client_id == client.id).all()
        }
        for pay in row.get("payments", []):
            try:
                pay_date = date.fromisoformat(pay["date"])
            except (KeyError, ValueError, TypeError):
                continue
            pay_amount = round(float(pay.get("amount") or 0), 2)
            if pay_amount <= 0:
                continue
            key = (pay_date, pay_amount)
            if key in existing_dates:
                real_payments_skipped += 1
                continue
            db.add(CollectionPayment(
                client_id=client.id,
                collection_date=pay_date,
                amount=pay_amount,
                payment_method="import",
                reference=f"Fiche collecte {filename or 'upload'}",
                notes="Importé depuis FICHE DE COLLECTE",
                created_at=_now(),
            ))
            existing_dates.add(key)
            real_payments_created += 1

        amount = row["typical_daily"] or row["montant_total"]
        if amount <= 0:
            skipped += 1
            errors.append(f"{name} : montant nul")
            continue

        label = f"Collecte — {name}"
        next_d: Optional[date] = None
        if row.get("last_date"):
            try:
                next_d = date.fromisoformat(row["last_date"])
            except ValueError:
                next_d = date.today()
        else:
            next_d = date.today()

        # PaymentRecurrence n'a qu'un champ texte libre (`description`, pas de colonne `notes`
        # séparée) — les deux informations sont donc combinées dedans plutôt que de planter
        # sur un attribut inexistant (bug confirmé : « type object 'PaymentRecurrence' has no
        # attribute 'notes' » sur tout import de fiche de collecte).
        description = (
            f"Import collecte {filename or 'upload'} · solde {row['montant_total']:.0f} · "
            f"{row['payments_count']} paiement(s)"
        )

        existing = db.query(PaymentRecurrence).filter(
            PaymentRecurrence.client_id == client.id,
            PaymentRecurrence.name == label,
            PaymentRecurrence.status != 'cancelled',
        ).first()

        if existing:
            existing.amount = amount
            existing.frequency_code = frequency
            existing.next_due_date = next_d
            existing.status = 'active'
            existing.is_active = True
            existing.description = description
            updated += 1
        else:
            db.add(PaymentRecurrence(
                company_id=company_id,
                created_by=0,
                updated_by=0,
                name=label,
                description=description,
                client_id=client.id,
                project_id=None,
                category='collecte',
                recurrence_type='collecte',
                amount=amount,
                currency_code='XOF',
                vat_rate=0,
                discount_pct=0,
                balance=0,
                frequency_code=frequency,
                start_date=date.today(),
                end_date=None,
                next_due_date=next_d,
                status='active',
                is_active=True,
                auto_generate=True,
                auto_notify=False,
                auto_followup=False,
                auto_invoice=False,
                draft_days_before=0,
            ))
            created += 1

    db.flush()
    msg = f"{real_payments_created} paiement(s) réel(s) importé(s), {created} plan(s) créé(s), {updated} mis à jour"
    if real_payments_skipped:
        msg += f", {real_payments_skipped} paiement(s) déjà présent(s)"
    if clients_created:
        msg += f", {clients_created} client(s) créé(s)"
    if skipped:
        msg += f", {skipped} ligne(s) ignorée(s)"

    return {
        "ok": True,
        "message": msg,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "clients_created": clients_created,
        "real_payments_created": real_payments_created,
        "real_payments_skipped": real_payments_skipped,
        "errors": errors,
        "parsed": {
            "total": parsed["total"],
            "dates": parsed["dates"],
            "filename": filename,
        },
    }
