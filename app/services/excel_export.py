"""Export Excel — états, listes, rapports (modèle complet, pas colonnes visibles UI)."""
from datetime import date, timedelta
from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import Client, Invoice, Transaction, StockItem, Account, User
from app.models_business_ext import InvoiceRecurring
from app.models_erp import Supplier, SupplierInvoice, JournalEntry, JournalLine, Employee
from app.services.accounting_reports import balance_generale, grand_livre

_RECURRING_FREQ_LBL = {
    "daily": "Journalière", "weekly": "Hebdomadaire", "monthly": "Mensuelle",
    "quarterly": "Trimestrielle", "yearly": "Annuelle",
}


def _recurring_monthly_equivalent(amount, frequency: str) -> float:
    amount = float(amount or 0)
    if frequency == "daily":
        return amount * 30
    if frequency == "weekly":
        return amount * 4.33
    if frequency == "quarterly":
        return amount / 3
    if frequency == "yearly":
        return amount / 12
    return amount


def _recurring_remaining(amount, frequency: str, next_date, end_date):
    """Reproduit exactement estimateRecurringRemaining() du frontend (business-suite-ui.js)."""
    if not end_date:
        return None
    amount = float(amount or 0)
    if not next_date or end_date < next_date:
        return amount
    months = max(0, (end_date.year - next_date.year) * 12 + (end_date.month - next_date.month))
    cycles = 1
    if frequency == "daily":
        cycles += max(1, (end_date - next_date).days)
    elif frequency == "weekly":
        cycles += int(months * 4.348)
    elif frequency == "monthly":
        cycles += months
    elif frequency == "quarterly":
        cycles += months // 3
    elif frequency == "yearly":
        cycles += months // 12
    return max(0.0, round(amount * cycles))


def _style_sheet(ws, currency_cols=None, date_cols=None):
    currency_cols = currency_cols or []
    date_cols = date_cols or []
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = __import__("openpyxl.styles", fromlist=["PatternFill"]).PatternFill(
            "solid", fgColor="5C4033"
        )
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin
            if cell.column in currency_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            if cell.column in date_cols and cell.value:
                cell.number_format = "DD/MM/YYYY"
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        maxlen = 0
        for cell in ws[letter]:
            maxlen = max(maxlen, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(42, max(12, maxlen + 2))
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _to_bytes(df: pd.DataFrame, sheet_name: str = "Données") -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name=sheet_name)
        _style_sheet(w.sheets[sheet_name])
    buf.seek(0)
    return buf.getvalue()


def _multi_sheet(sheets: dict[str, pd.DataFrame], currency_by_sheet=None, date_by_sheet=None) -> bytes:
    currency_by_sheet = currency_by_sheet or {}
    date_by_sheet = date_by_sheet or {}
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        for name, df in sheets.items():
            safe = (name or "Feuille")[:31]
            df.to_excel(w, index=False, sheet_name=safe)
            ws = w.sheets[safe]
            headers = list(df.columns)
            cur_idx = [headers.index(c) + 1 for c in currency_by_sheet.get(name, []) if c in headers]
            date_idx = [headers.index(c) + 1 for c in date_by_sheet.get(name, []) if c in headers]
            _style_sheet(ws, currency_cols=cur_idx, date_cols=date_idx)
    buf.seek(0)
    return buf.getvalue()


def export_clients(db: Session) -> bytes:
    """Export complet — tous les champs client, avec des en-têtes reconnus par le
    ré-import (excel_mapper/COLUMN_ALIASES) pour permettre modifier puis ré-importer
    sans dupliquer (upsert par « ID Client », voir imports_attachments._import_client_row)."""
    rows = []
    for c in db.query(Client).all():
        rows.append({
            "ID Client": c.id,
            "Nom": c.name or "",
            "Type": c.type or "",
            "Segment": c.segment or "",
            "Contact": c.contact or "",
            "Email": c.email or "",
            "Téléphone": c.phone or "",
            "Ville": c.city or "",
            "Statut": c.status or "",
            "Compte": c.account_code or "",
            "Limite crédit": float(c.credit_limit or 0),
            "Délai paiement": int(c.payment_terms or 0),
            "Remise": float(c.default_discount_pct or 0),
            "TVA": float(c.default_vat_pct or 0),
            "Commission": float(c.default_commission_pct or 0),
            "Retenue source": float(c.default_withholding_pct or 0),
            "Archivé": "Oui" if c.is_archived else "Non",
            "Notes": c.notes or "",
        })
    df = pd.DataFrame(rows or [{"info": "aucun client"}])
    return _multi_sheet({"Clients": df}, currency_by_sheet={"Clients": ["Limite crédit"]})


def export_invoices(db: Session) -> bytes:
    """Export complet factures : en-têtes + lignes + synthèse (HT/TVA/TTC/soldes)."""
    from app.models import DocumentLine, Quote
    from app.services.document_calc_service import compute_document_totals

    clients = {c.id: c for c in db.query(Client).all()}
    users = {u.id: (u.full_name or u.email or str(u.id)) for u in db.query(User).all()}
    invoices = db.query(Invoice).order_by(Invoice.date.desc(), Invoice.id.desc()).all()

    header_rows = []
    line_rows = []
    total_ht = total_tva = total_ttc = total_paid = total_reste = 0.0

    for inv in invoices:
        cl = clients.get(inv.client_id)
        lines = (
            db.query(DocumentLine)
            .filter(DocumentLine.invoice_id == inv.id)
            .order_by(DocumentLine.position, DocumentLine.id)
            .all()
        )
        totals = compute_document_totals(
            db,
            lines,
            doc_vat_rate=getattr(inv, "vat_rate", None),
            discount_pct=getattr(inv, "discount_pct", 0) or 0,
            commission_pct=getattr(inv, "commission_pct", 0) or 0,
            paid=getattr(inv, "paid", 0) or 0,
        ) if lines else {
            "amount_ht": float(inv.amount or 0),
            "total_tva": 0.0,
            "amount_ttc": float(inv.amount or 0),
            "subtotal_ht": float(inv.amount or 0),
            "commission": 0.0,
            "global_disc": 0.0,
            "paid": float(inv.paid or 0),
            "remaining": max(0.0, float(inv.amount or 0) - float(inv.paid or 0)),
        }
        # Si pas de lignes, estimer TVA depuis vat_rate document
        if not lines and float(getattr(inv, "vat_rate", 0) or 0) > 0:
            ht = float(inv.amount or 0)
            rate = float(inv.vat_rate or 0)
            tva = round(ht * rate / 100)
            totals = {
                **totals,
                "amount_ht": ht,
                "total_tva": tva,
                "amount_ttc": ht + tva,
                "remaining": max(0.0, ht + tva - float(inv.paid or 0)),
            }

        ht = float(totals.get("amount_ht") or 0)
        tva = float(totals.get("total_tva") or 0)
        ttc = float(totals.get("amount_ttc") or ht)
        paid = float(inv.paid or 0)
        reste = max(0.0, ttc - paid) if lines else max(0.0, float(inv.amount or 0) - paid)

        total_ht += ht
        total_tva += tva
        total_ttc += ttc
        total_paid += paid
        total_reste += reste

        header_rows.append({
            "id": inv.id,
            "numero": inv.number or "",
            "type_document": getattr(inv, "doc_type", None) or "invoice",
            "date": inv.date.isoformat() if inv.date else "",
            "date_echeance": inv.due_date.isoformat() if inv.due_date else "",
            "date_livraison": inv.delivery_date.isoformat() if getattr(inv, "delivery_date", None) else "",
            "client": cl.name if cl else "",
            "code_client": inv.client_id or "",
            "compte_client": (cl.account_code if cl else "") or "",
            "email_client": (cl.email if cl else "") or "",
            "telephone_client": (cl.phone if cl else "") or "",
            "montant_ht": ht,
            "remise_pct": float(getattr(inv, "discount_pct", 0) or 0),
            "commission_pct": float(getattr(inv, "commission_pct", 0) or 0),
            "retenue_pct": float(getattr(inv, "withholding_pct", 0) or 0),
            "taux_tva": float(getattr(inv, "vat_rate", 0) or 0),
            "montant_tva": tva,
            "montant_ttc": ttc,
            "paye": paid,
            "reste_a_payer": reste,
            "devise": "XOF",
            "statut": inv.status or "",
            "conditions_paiement_jours": getattr(inv, "payment_terms_days", None) or "",
            "devis_lie_id": getattr(inv, "quote_id", None) or "",
            "plan_recurrent_id": getattr(inv, "recurring_plan_id", None) or "",
            "projet_id": getattr(inv, "project_id", None) or "",
            "notes": getattr(inv, "notes", None) or "",
            "description": getattr(inv, "description", None) or "",
            "cree_le": getattr(inv, "created_at", None) or "",
            "cree_par": users.get(getattr(inv, "created_by", None), ""),
            "nb_lignes": len(lines),
        })

        for ln in lines:
            qte = float(ln.quantity or 0)
            pu = float(ln.unit_price or 0)
            disc = float(getattr(ln, "discount_pct", 0) or 0)
            brut = round(qte * pu)
            ht_l = round(brut * (1 - min(100, max(0, disc)) / 100)) if disc else brut
            vat_l = float(ln.vat_rate or 0)
            tva_l = round(ht_l * vat_l / 100)
            line_rows.append({
                "facture_id": inv.id,
                "numero_facture": inv.number or "",
                "client": cl.name if cl else "",
                "position": ln.position or 0,
                "reference": getattr(ln, "reference", None) or "",
                "designation": ln.description or "",
                "unite": getattr(ln, "unit", None) or "unité",
                "quantite": qte,
                "prix_unitaire_ht": pu,
                "remise_pct": disc,
                "montant_ht": ht_l,
                "taux_tva": vat_l,
                "montant_tva": tva_l,
                "montant_ttc": ht_l + tva_l,
                "compte": getattr(ln, "account_code", None) or getattr(ln, "sale_account", None) or "",
                "type_produit": getattr(ln, "product_type", None) or "",
            })

    data_df = pd.DataFrame(header_rows or [{"id": "", "numero": "", "info": "aucune facture"}])
    lines_df = pd.DataFrame(line_rows or [{"facture_id": "", "designation": "", "info": "aucune ligne"}])
    synth = pd.DataFrame([
        {"indicateur": "Nombre de factures", "valeur": len(invoices)},
        {"indicateur": "Nombre de lignes", "valeur": len(line_rows)},
        {"indicateur": "Total HT", "valeur": round(total_ht, 2)},
        {"indicateur": "Total TVA", "valeur": round(total_tva, 2)},
        {"indicateur": "Total TTC", "valeur": round(total_ttc, 2)},
        {"indicateur": "Total encaissé", "valeur": round(total_paid, 2)},
        {"indicateur": "Total reste à payer", "valeur": round(total_reste, 2)},
        {"indicateur": "Devise", "valeur": "XOF"},
    ])

    return _multi_sheet(
        {"Factures": data_df, "Lignes": lines_df, "Synthèse": synth},
        currency_by_sheet={
            "Factures": ["montant_ht", "montant_tva", "montant_ttc", "paye", "reste_a_payer"],
            "Lignes": ["prix_unitaire_ht", "montant_ht", "montant_tva", "montant_ttc"],
            "Synthèse": ["valeur"],
        },
        date_by_sheet={"Factures": ["date", "date_echeance", "date_livraison"]},
    )


def export_quotes(db: Session) -> bytes:
    """Export complet devis : en-têtes + lignes + synthèse."""
    from app.models import DocumentLine, Quote
    from app.services.document_calc_service import compute_document_totals

    clients = {c.id: c for c in db.query(Client).all()}
    quotes = db.query(Quote).order_by(Quote.date.desc(), Quote.id.desc()).all()

    header_rows = []
    line_rows = []
    total_ht = total_tva = total_ttc = 0.0

    for q in quotes:
        cl = clients.get(q.client_id)
        lines = (
            db.query(DocumentLine)
            .filter(DocumentLine.quote_id == q.id)
            .order_by(DocumentLine.position, DocumentLine.id)
            .all()
        )
        if lines:
            totals = compute_document_totals(
                db, lines, doc_vat_rate=getattr(q, "vat_rate", None),
            )
        else:
            ht = float(q.amount or 0)
            rate = float(getattr(q, "vat_rate", 0) or 0)
            tva = round(ht * rate / 100) if rate else 0
            totals = {"amount_ht": ht, "total_tva": tva, "amount_ttc": ht + tva}

        ht = float(totals.get("amount_ht") or 0)
        tva = float(totals.get("total_tva") or 0)
        ttc = float(totals.get("amount_ttc") or ht)
        total_ht += ht
        total_tva += tva
        total_ttc += ttc

        header_rows.append({
            "id": q.id,
            "numero": q.number or "",
            "objet": q.title or "",
            "date": q.date.isoformat() if q.date else "",
            "validite": q.valid_until.isoformat() if q.valid_until else "",
            "client": cl.name if cl else "",
            "code_client": q.client_id or "",
            "compte_client": (cl.account_code if cl else "") or "",
            "email_client": (cl.email if cl else "") or "",
            "telephone_client": (cl.phone if cl else "") or "",
            "montant_ht": ht,
            "taux_tva": float(getattr(q, "vat_rate", 0) or 0),
            "montant_tva": tva,
            "montant_ttc": ttc,
            "devise": "XOF",
            "statut": q.status or "",
            "deal_id": getattr(q, "deal_id", None) or "",
            "notes": getattr(q, "notes", None) or "",
            "nb_lignes": len(lines),
        })

        for ln in lines:
            qte = float(ln.quantity or 0)
            pu = float(ln.unit_price or 0)
            disc = float(getattr(ln, "discount_pct", 0) or 0)
            brut = round(qte * pu)
            ht_l = round(brut * (1 - min(100, max(0, disc)) / 100)) if disc else brut
            vat_l = float(ln.vat_rate or 0)
            tva_l = round(ht_l * vat_l / 100)
            line_rows.append({
                "devis_id": q.id,
                "numero_devis": q.number or "",
                "client": cl.name if cl else "",
                "position": ln.position or 0,
                "reference": getattr(ln, "reference", None) or "",
                "designation": ln.description or "",
                "unite": getattr(ln, "unit", None) or "unité",
                "quantite": qte,
                "prix_unitaire_ht": pu,
                "remise_pct": disc,
                "montant_ht": ht_l,
                "taux_tva": vat_l,
                "montant_tva": tva_l,
                "montant_ttc": ht_l + tva_l,
                "compte": getattr(ln, "account_code", None) or "",
            })

    data_df = pd.DataFrame(header_rows or [{"id": "", "numero": "", "info": "aucun devis"}])
    lines_df = pd.DataFrame(line_rows or [{"devis_id": "", "designation": "", "info": "aucune ligne"}])
    synth = pd.DataFrame([
        {"indicateur": "Nombre de devis", "valeur": len(quotes)},
        {"indicateur": "Nombre de lignes", "valeur": len(line_rows)},
        {"indicateur": "Total HT", "valeur": round(total_ht, 2)},
        {"indicateur": "Total TVA", "valeur": round(total_tva, 2)},
        {"indicateur": "Total TTC", "valeur": round(total_ttc, 2)},
        {"indicateur": "Devise", "valeur": "XOF"},
    ])

    return _multi_sheet(
        {"Devis": data_df, "Lignes": lines_df, "Synthèse": synth},
        currency_by_sheet={
            "Devis": ["montant_ht", "montant_tva", "montant_ttc"],
            "Lignes": ["prix_unitaire_ht", "montant_ht", "montant_tva", "montant_ttc"],
            "Synthèse": ["valeur"],
        },
        date_by_sheet={"Devis": ["date", "validite"]},
    )


def export_transactions(db: Session) -> bytes:
    rows = [{"date": str(t.date), "libelle": t.label, "type": t.type, "montant": t.amount, "categorie": t.category}
            for t in db.query(Transaction).all()]
    return _to_bytes(pd.DataFrame(rows or [{"info": "aucune ecriture"}]), "Ecritures")


def export_journal_entries(db: Session) -> bytes:
    """Export full-model des écritures SYSCOHADA + feuille synthèse."""
    clients = {c.id: c for c in db.query(Client).all()}
    users = {u.id: (u.full_name or u.email or str(u.id)) for u in db.query(User).all()}
    accounts = {a.code: a.label for a in db.query(Account).all()} if hasattr(Account, "code") else {}
    try:
        from app.syscohada.chart import get_account_label
    except Exception:
        get_account_label = None

    rows = []
    total_debit = 0.0
    total_credit = 0.0
    entries = db.query(JournalEntry).order_by(JournalEntry.date.desc(), JournalEntry.id.desc()).all()
    for e in entries:
        cl = clients.get(getattr(e, "client_id", None))
        lines = db.query(JournalLine).filter(JournalLine.entry_id == e.id).order_by(JournalLine.id).all()
        for ln in lines:
            line_cl = clients.get(getattr(ln, "client_id", None)) or cl
            acc_label = accounts.get(ln.account_code) or ""
            if not acc_label and get_account_label:
                try:
                    acc_label = get_account_label(ln.account_code) or ""
                except Exception:
                    acc_label = ""
            debit = float(ln.debit or 0)
            credit = float(ln.credit or 0)
            total_debit += debit
            total_credit += credit
            rows.append({
                "id": e.id,
                "date": e.date.isoformat() if e.date else "",
                "numero_piece": e.reference or f"JE-{e.id}",
                "type_document": e.source_type or "journal",
                "journal": e.journal or "",
                "numero_compte": ln.account_code,
                "libelle_compte": acc_label,
                "nom_client": (line_cl.name if line_cl else "") or "",
                "code_client": (line_cl.id if line_cl else "") or "",
                "reference_facture": e.reference or "",
                "libelle": ln.label or e.label or "",
                "montant_debit": debit,
                "montant_credit": credit,
                "devise": "XOF",
                "statut": e.status or "",
                "date_creation": getattr(e, "created_at", "") or "",
                "utilisateur": users.get(getattr(e, "created_by", None), ""),
                "observations": getattr(e, "observation", "") or "",
            })

    data_df = pd.DataFrame(rows or [{
        "id": "", "date": "", "numero_piece": "", "type_document": "", "journal": "",
        "numero_compte": "", "libelle_compte": "", "nom_client": "", "code_client": "",
        "reference_facture": "", "libelle": "", "montant_debit": 0, "montant_credit": 0,
        "devise": "XOF", "statut": "", "date_creation": "", "utilisateur": "", "observations": "",
    }])

    synth = pd.DataFrame([
        {"indicateur": "Nombre d'écritures", "valeur": len(entries)},
        {"indicateur": "Nombre de lignes", "valeur": len(rows)},
        {"indicateur": "Total débit", "valeur": round(total_debit, 2)},
        {"indicateur": "Total crédit", "valeur": round(total_credit, 2)},
        {"indicateur": "Solde (débit − crédit)", "valeur": round(total_debit - total_credit, 2)},
        {"indicateur": "Devise", "valeur": "XOF"},
    ])

    by_journal = {}
    for r in rows:
        j = r.get("journal") or "—"
        by_journal.setdefault(j, {"journal": j, "debit": 0.0, "credit": 0.0, "lignes": 0})
        by_journal[j]["debit"] += r["montant_debit"]
        by_journal[j]["credit"] += r["montant_credit"]
        by_journal[j]["lignes"] += 1
    journals_df = pd.DataFrame(list(by_journal.values()) or [{"journal": "—", "debit": 0, "credit": 0, "lignes": 0}])

    return _multi_sheet(
        {"Données": data_df, "Synthèse": synth, "Par journal": journals_df},
        currency_by_sheet={
            "Données": ["montant_debit", "montant_credit"],
            "Synthèse": ["valeur"],
            "Par journal": ["debit", "credit"],
        },
        date_by_sheet={"Données": ["date"]},
    )


def export_balance(db: Session) -> bytes:
    return _to_bytes(pd.DataFrame(balance_generale(db)), "Balance")


def export_grand_livre(db: Session) -> bytes:
    return _to_bytes(pd.DataFrame(grand_livre(db)), "Grand livre")


def export_stock(db: Session) -> bytes:
    rows = [{"sku": s.sku, "nom": s.name, "quantite": s.quantity, "seuil": s.min_quantity, "valeur": s.quantity * s.unit_cost}
            for s in db.query(StockItem).all()]
    return _to_bytes(pd.DataFrame(rows), "Stock")


def export_suppliers(db: Session) -> bytes:
    rows = [{"nom": s.name, "email": s.email, "ville": s.city, "statut": s.status} for s in db.query(Supplier).all()]
    return _to_bytes(pd.DataFrame(rows or [{"info": "aucun"}]), "Fournisseurs")


def export_employees(db: Session) -> bytes:
    rows = [{"matricule": e.matricule, "nom": f"{e.firstname} {e.lastname}", "dept": e.department, "salaire": e.salary_base}
            for e in db.query(Employee).all()]
    return _to_bytes(pd.DataFrame(rows or [{"info": "aucun"}]), "Employes")


def export_recurring(db: Session) -> bytes:
    """Export complet des plans de facturation récurrente (détail + synthèse chiffrée)."""
    clients = {c.id: c for c in db.query(Client).all()}
    plans = db.query(InvoiceRecurring).order_by(InvoiceRecurring.id.desc()).all()
    today = date.today()
    in7 = today + timedelta(days=7)

    rows = []
    total_amount_active = 0.0
    total_monthly_active = 0.0
    total_remaining = 0.0
    n_active = n_inactive = n_overdue = n_due_soon = n_with_end_date = 0

    for p in plans:
        cl = clients.get(p.client_id)
        active = bool(p.active)
        overdue = active and p.next_date is not None and p.next_date < today
        due_soon = active and not overdue and p.next_date is not None and p.next_date <= in7
        if active:
            n_active += 1
            total_amount_active += float(p.amount or 0)
            total_monthly_active += _recurring_monthly_equivalent(p.amount, p.frequency)
        else:
            n_inactive += 1
        n_overdue += 1 if overdue else 0
        n_due_soon += 1 if due_soon else 0

        remaining = _recurring_remaining(p.amount, p.frequency, p.next_date, p.end_date)
        if remaining is not None:
            n_with_end_date += 1
            total_remaining += remaining

        statut = "Suspendu" if not active else ("En retard" if overdue else ("Échéance proche" if due_soon else "Actif"))

        rows.append({
            "id": p.id,
            "libelle": p.label or "",
            "client": cl.name if cl else "",
            "code_client": p.client_id,
            "compte_client": (cl.account_code if cl else "") or "",
            "frequence": _RECURRING_FREQ_LBL.get(p.frequency, p.frequency or ""),
            "prochaine_echeance": str(p.next_date) if p.next_date else "",
            "date_fin": str(p.end_date) if p.end_date else "",
            "montant_echeance": float(p.amount or 0),
            "equivalent_mensuel": round(_recurring_monthly_equivalent(p.amount, p.frequency), 2),
            "reste_a_facturer": remaining if remaining is not None else "",
            "statut": statut,
            "derniere_generation": p.last_generated_at or "",
            "facture_modele_id": p.template_invoice_id or "",
            "notes": p.notes or "",
        })

    data_df = pd.DataFrame(rows or [{
        "id": "", "libelle": "", "client": "", "code_client": "", "compte_client": "",
        "frequence": "", "prochaine_echeance": "", "date_fin": "", "montant_echeance": 0,
        "equivalent_mensuel": 0, "reste_a_facturer": "", "statut": "", "derniere_generation": "",
        "facture_modele_id": "", "notes": "",
    }])

    synth = pd.DataFrame([
        {"indicateur": "Nombre total de plans", "valeur": len(plans)},
        {"indicateur": "Plans actifs", "valeur": n_active},
        {"indicateur": "Plans suspendus", "valeur": n_inactive},
        {"indicateur": "Échéances en retard", "valeur": n_overdue},
        {"indicateur": "Échéances sous 7 jours", "valeur": n_due_soon},
        {"indicateur": "Montant total / échéance (plans actifs)", "valeur": round(total_amount_active, 2)},
        {"indicateur": "Revenu récurrent mensuel (équivalent, actifs)", "valeur": round(total_monthly_active, 2)},
        {"indicateur": "Revenu récurrent annuel (projection)", "valeur": round(total_monthly_active * 12, 2)},
        {"indicateur": "Reste à facturer (plans à durée déterminée)", "valeur": round(total_remaining, 2)},
        {"indicateur": "Plans avec date de fin", "valeur": f"{n_with_end_date}/{len(plans)}"},
        {"indicateur": "Devise", "valeur": "XOF"},
    ])

    return _multi_sheet(
        {"Plans récurrents": data_df, "Synthèse": synth},
        currency_by_sheet={
            "Plans récurrents": ["montant_echeance", "equivalent_mensuel", "reste_a_facturer"],
            "Synthèse": ["valeur"],
        },
        date_by_sheet={"Plans récurrents": ["prochaine_echeance", "date_fin"]},
    )
